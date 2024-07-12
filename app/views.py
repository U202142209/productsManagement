import ast
import json

import jwt
import traceback

from django.http import HttpResponse
from django.shortcuts import render
from django.db.models import Q, Min
from django.forms import model_to_dict
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.conf import settings
from openpyxl.workbook import Workbook
from rest_framework.response import Response
from rest_framework.views import APIView

TOKEN_secret_key = settings.TOKEN_SECRET_KEY
from .models import LyqInfo, ApInfo, JhjInfo, SxtInfo, FucaiInfo, NVRInfo, YinPanInfo
from .models import Group, Quyu, User, Management
from .utils.SetJsonResponse import setSuccessResponse, setErrorResponse, setMyJson, setMyError


# http://127.0.0.1:8000/api/submit


@require_POST  # 仅限于post请求
@csrf_exempt  # 去除csrf token认证
def submit(request):
    try:
        # 验证token
        authorization_header = request.META.get('HTTP_AUTHORIZATION')
        token = str(authorization_header).split(" ")[1]
        # 验证token
        payload = jwt.decode(token, TOKEN_secret_key, algorithms=['HS256'])
        phone_number = payload["phone_number"]
        if not User.objects.filter(phone_number=phone_number).exists():
            return setErrorResponse("请先登录")
        # 解析参数

        ap_number = request.POST.get("ap_number")  # ap 数量
        ap_factory = request.POST.get("ap_factory", "")  # ap 品牌
        ap_classify = request.POST.get("ap_classify")  # ap 分类
        ap_color = request.POST.get("ap_color", "")  # 颜色
        ap_wifi_sulv = request.POST.get("wifi_sulv")  # wifi速率
        ap_user_number = request.POST.get("user_number", "")  # 用户数
        sxt_number = request.POST.get("sxt_number", 0)  # 摄像头数量
        sxt_factory = request.POST.get("sxt_factory", 1)  # sxt 品牌
        sxt_classify = request.POST.get("sxt_classify", 1)  # sxt 分类
        sxt_px = request.POST.get("sxt_px", 400)  # sxt 像素
        sxt_jiaoju = request.POST.get("sxt_jiaoju", "4mm")
        sxt_yeshi = request.POST.get("sxt_yeshi", 1)  # sxt 夜视
        sxt_ccfs = request.POST.get("sxt_ccfs", 1)  # sxt 存储方式,硬盘还是SD卡
        sxt_ccsc = request.POST.get("sxt_ccsc", 1)  # sxt 存储时长/天,是用来计算硬盘大小

        # 根据参数写查询逻辑
        result_ap = ApInfo.objects.filter().first()
        if not result_ap:
            return setErrorResponse("没有查询到数据")
        return setSuccessResponse(
            msg="ok",
            data={
                "数据": [
                    result_ap.to_dict(number=4),
                    result_ap.to_dict(number=4),
                    result_ap.to_dict(number=4),
                    result_ap.to_dict(number=4),
                ],
                "价格": {}
            }
        )
    except jwt.ExpiredSignatureError:
        return setErrorResponse(msg="Token expired", code=1002)
    except jwt.InvalidTokenError:
        return setErrorResponse(msg="Invalid token", code=1003)
    except IndexError:
        return setErrorResponse(msg="token is needed", code=1001)
    except Exception as e:
        traceback.print_exc()
        return setErrorResponse(msg="server error")


@require_POST  # 仅限于post请求
@csrf_exempt  # 去除csrf token认证
def submit_fangan(request):
    try:
        # 验证token
        authorization_header = request.META.get('HTTP_AUTHORIZATION')
        token = str(authorization_header).split(" ")[1]
        payload = jwt.decode(token, TOKEN_secret_key, algorithms=['HS256'])
        phone_number = payload["phone_number"]
        u = User.objects.filter(phone_number=phone_number).first()
        if not u:
            return setErrorResponse("请先登录")
        # 获取参数
        name = request.POST.get("name", "")
        if not name:
            return setErrorResponse("请输入方案名称")
        content = request.POST.get("content", "")
        if not content:
            return setErrorResponse("请输入方案内容")
        # 开始保存方案
        Management.objects.create(
            content=ast.literal_eval(content),
            user=u,
            name=name,
            status=1
        )
        return setSuccessResponse("方案提交成功")
    except jwt.ExpiredSignatureError:
        return setErrorResponse(msg="Token expired", code=1002)
    except jwt.InvalidTokenError:
        return setErrorResponse(msg="Invalid token", code=1003)
    except IndexError:
        return setErrorResponse(msg="token is needed", code=1001)
    except Exception as e:
        traceback.print_exc()
        return setErrorResponse(msg="server error")


# 导出为excel
def export_as_excel(request):
    id = request.GET.get("id", "")
    m = Management.objects.filter(id=id).first()
    if not m:
        return setErrorResponse("id不正确")
    content = m.content
    # 导出为excel
    # 创建一个新的Excel工作簿
    wb = Workbook()
    ws = wb.active
    # 写入表头
    header_row = ['序号', '设备类型', '型号', '参数概要', '数量', '单位', '设备单价', '设备小计',
                  '施工费', '备注']
    for col_num, header in enumerate(header_row, 1):
        ws.cell(row=1, column=col_num, value=header)
    # 将模型中的数据写入工作表
    for row_num, item in enumerate(content["数据"], 2):
        ws.cell(row=row_num, column=1, value=row_num - 1)
        ws.cell(row=row_num, column=2, value=item.get("设备类型", ""))
        ws.cell(row=row_num, column=3, value=item.get("型号", ""))
        ws.cell(row=row_num, column=4, value=item.get("参数概要", ""))
        ws.cell(row=row_num, column=5, value=item.get("数量", ""))
        ws.cell(row=row_num, column=6, value=item.get("单位", ""))
        ws.cell(row=row_num, column=7, value=item.get("设备单价", ""))
        ws.cell(row=row_num, column=8, value=item.get("设备小计", ""))
        ws.cell(row=row_num, column=9, value=item.get("施工费", ""))
        ws.cell(row=row_num, column=10, value=item.get("备注", ""))
    # 设置响应头，指定文件名和文件类型
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=用户信息.xlsx'
    # 将Excel内容写入响应
    wb.save(response)
    return response


# 纯AP方案
@require_POST  # 仅限于post请求
@csrf_exempt  # 去除csrf token认证
def api_c_ap(request):
    try:
        print(request.POST)
        # ap 数量
        ap_number = request.POST.get("ap_number", 0)
        print(ap_number)
        # ap 品牌
        ap_factory = request.POST.get("ap_factory", 1)
        # ap 分类
        ap_classify = request.POST.get("ap_classify", 1)
        # 颜色
        color = request.POST.get("ap_color", "白色")
        # wifi速率
        wifi_sulv = request.POST.get("wifi_sulv", 1500)
        # 用户数
        user_number = request.POST.get("user_number", 100)
        # 查询ap
        aps = ApInfo.objects.filter(
            # 查询品牌商的名称 包含关键字
            factory__name__icontains=ap_factory,
            classify=int(ap_classify),
            color__icontains=color,
            lianjie_sulv=int(wifi_sulv),
        ).values("product_kind", "xiaoshoujia")
        if not aps:
            return setErrorResponse("没有查询到ap")
        # ......
        sxt_number = request.POST.get("sxt_number", 0)

        # 校验参数是否合法
        if sxt_number == 0:
            # 纯 ap 数
            if 1 <= int(ap_number) <= 4:
                # 默认5口一体化
                # 大于 gt  小于： lt 大于等于:gte
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(1, 4)).values("product_kind", "xiaoshoujia")
                # 查询所有字段
                # lyq = LyqInfo.objects.filter(
                #      classify=1, POE__lte=ap_number)
                if not lyq:
                    return setErrorResponse("没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        # 查询所有字段,输出
                        # "lyq": [model_to_dict(i) for i in ls],
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )
            elif 5 <= int(ap_number) <= 8:
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(5, 8)).values("product_kind", "xiaoshoujia")
                if not lyq:
                    return setErrorResponse("AP没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )
            elif 9 <= int(ap_number) <= 12:
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(5, 8)).values("product_kind", "xiaoshoujia")
                if not lyq:
                    return setErrorResponse("路由器没有查询到数据")

                jhj = JhjInfo.objects.filter(
                    classify=4, poe_number__range=(1, 4)).values("product_kind", "xiaoshoujia")

                if not jhj:
                    return setErrorResponse("交换机没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "jhj": [i for i in jhj],
                        "jhj_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )
            elif 13 <= int(ap_number) <= 16:
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(5, 8)).values("product_kind", "xiaoshoujia")
                if not lyq:
                    return setErrorResponse("路由器没有查询到数据")

                jhj = JhjInfo.objects.filter(
                    # 交换出来有三款,判断价格最便宜一款(不会写)
                    classify=4, poe_number__range=(5, 8), upstream_number__gte=1).values("product_kind", "xiaoshoujia")

                if not jhj:
                    return setErrorResponse("交换机没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "jhj": [i for i in jhj],
                        "jhj_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )
            else:
                return setSuccessResponse(
                    msg="逻辑复杂,脑细包烧光了",
                    data={
                        "提示": "请记录方案需求,联系17366909676",
                        "原因": "ap 数量大于等于13",
                    }
                )

        else:
            return setErrorResponse("摄像头数量不是0")

    except Exception as error:
        traceback.print_exc()
        return setErrorResponse("出错了:" + str(error))


# 纯摄像头方案
@require_POST  # 仅限于post请求
@csrf_exempt  # 去除csrf token认证
def api_c_sxt(request):
    try:
        print(request.POST)
        # sxt 数量
        sxt_number = request.POST.get("sxt_number", 0)
        print(sxt_number)
        # sxt 品牌
        sxt_factory = request.POST.get("sxt_factory", 1)
        # sxt 分类
        sxt_classify = request.POST.get("sxt_classify", 1)
        print("sxt_classify", sxt_classify)
        # sxt 夜视
        sxt_yeshi = request.POST.get("sxt_yeshi", 1)
        # sxt 像素
        sxt_px = request.POST.get("sxt_px", 400)
        #  sxt 存储方式,硬盘还是SD卡
        sxt_ccfs = request.POST.get("sxt_ccfs", 1)
        # sxt 存储时长/天,是用来计算硬盘大小
        sxt_ccsc = request.POST.get("sxt_ccsc", 1)
        # ap 数量
        ap_number = request.POST.get("ap_number", 0)
        # ap 品牌
        ap_factory = request.POST.get("ap_factory", 1)
        # ap 分类
        ap_classify = request.POST.get("ap_classify", 1)
        # 颜色
        color = request.POST.get("ap_color", "白色")
        # wifi速率
        wifi_sulv = request.POST.get("wifi_sulv", 1500)
        # 用户数
        user_number = request.POST.get("user_number", 100)
        # 查询sxt
        sxt = SxtInfo.objects.filter(
            # 查询品牌 分类 夜视 像素
            factory__name__icontains=sxt_factory,
            classify=int(sxt_classify),
            yeshi=sxt_yeshi,
            px=int(sxt_px),
        ).values("product_kind", "xiaoshoujia")
        if not sxt:
            return setErrorResponse("没有查询到摄像头")
        # 查询ap
        aps = ApInfo.objects.filter(
            # 查询品牌商的名称 包含关键字
            factory__name__icontains=ap_factory,
            classify=int(ap_classify),
            color__icontains=color,
            lianjie_sulv=int(wifi_sulv),
        ).values("product_kind", "xiaoshoujia")

        # 校验参数是否合法,ap为0
        if ap_number == 0:
            # 纯 ap 数
            if 1 <= int(sxt_number) <= 4:
                # 默认百兆POE交换机
                # 大于 gt  小于： lt 大于等于:gte
                jhj = JhjInfo.objects.filter(
                    classify=3, poe_number__range=(1, 4), upstream_number__gte=1).values("product_kind", "xiaoshoujia")

                if not jhj:
                    return setErrorResponse("4口百兆POE交换机没有查询到数据")

                fc_zj = FucaiInfo.objects.filter(
                    classify_fl="枪机支架"
                ).values("product_kind", "xiaoshoujia")
                if not fc_zj:
                    return setErrorResponse("辅材枪机支架没有查询到数据")

                nvr = NVRInfo.objects.filter(
                    jierunshu__range=(1, 4),
                ).values("product_kind", "xiaoshoujia")
                if not nvr:
                    return setErrorResponse("NVR录像机没有查询到数据")
                ypdx = YinPanInfo.objects.filter(
                    classify=1,
                ).values("product_kind", "xiaoshoujia")
                if not ypdx:
                    return setErrorResponse("硬盘没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "sxt": [i for i in sxt],
                        "sxt_sl": sxt_number,
                        "fc_zj": [i for i in fc_zj],
                        "fc_zj_sl": sxt_number,
                        "jhj": [i for i in jhj],
                        "jhj_sl": 1,
                        "nvr": [i for i in nvr],
                        "nvr_sl": 1,
                        "ypdx": [i for i in ypdx],
                        "ypdx_sl": 1,
                    }
                )
            elif 5 <= int(sxt_number) <= 8:
                jhj = JhjInfo.objects.filter(
                    classify=3, poe_number__range=(5, 8), upstream_number__gte=1).values("product_kind", "xiaoshoujia")

                if not jhj:
                    return setErrorResponse("8口百兆POE交换机没有查询到数据,注意查询参数有上联口为>1")
                if not jhj:
                    return setErrorResponse("4口百兆POE交换机没有查询到数据")
                fc_zj = FucaiInfo.objects.filter(
                    classify_fl="枪机支架"
                ).values("product_kind", "xiaoshoujia")
                if not fc_zj:
                    return setErrorResponse("辅材枪机支架没有查询到数据")

                nvr = NVRInfo.objects.filter(
                    jierunshu__range=(5, 8),
                ).values("product_kind", "xiaoshoujia")
                if not nvr:
                    return setErrorResponse("NVR录像机没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "sxt": [i for i in sxt],
                        "sxt_sl": sxt_number,
                        "fc_zj": [i for i in fc_zj],
                        "fc_zj_sl": sxt_number,
                        "jhj": [i for i in jhj],
                        "jhj_sl": 1,
                        "nvr": [i for i in nvr],
                        "nvr_sl": 1,
                    }
                )
            elif 9 <= int(sxt_number) <= 16:
                # 出来了两款,按价格低的优先(不会)
                jhj = JhjInfo.objects.filter(
                    classify=4, poe_number__range=(9, 16),
                    upstream_number__gte=1
                ).values("product_kind", "xiaoshoujia").order_by("xiaoshoujia").first()
                if not jhj:
                    return setErrorResponse("16口千兆POE交换机没有查询到数据,注意查询参数有上联口为>1")
                fc_zj = FucaiInfo.objects.filter(
                    classify_fl="枪机支架"
                ).values("product_kind", "xiaoshoujia")
                if not fc_zj:
                    return setErrorResponse("辅材枪机支架没有查询到数据")

                nvr = NVRInfo.objects.filter(
                    jierunshu__range=(9, 16),
                ).values("product_kind", "xiaoshoujia")
                if not nvr:
                    return setErrorResponse("NVR录像机没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "sxt": [i for i in sxt],
                        "sxt_sl": sxt_number,
                        "fc_zj": [i for i in fc_zj],
                        "fc_zj_sl": sxt_number,
                        "jhj": jhj,
                        "jhj_sl": 1,
                        "nvr": [i for i in nvr],
                        "nvr_sl": 1,
                    }
                )
            elif 17 <= int(sxt_number) <= 23:
                jhj = JhjInfo.objects.filter(
                    classify=4, poe_number__range=(17, 24)
                ).values("product_kind", "xiaoshoujia")
                if not jhj:
                    return setErrorResponse("24口千兆POE交换机没有查询到数据")
                fc_zj = FucaiInfo.objects.filter(
                    classify_fl="枪机支架"
                ).values("product_kind", "xiaoshoujia")
                if not fc_zj:
                    return setErrorResponse("辅材枪机支架没有查询到数据")

                nvr = NVRInfo.objects.filter(
                    jierunshu__range=(16, 32),
                ).values("product_kind", "xiaoshoujia")
                if not nvr:
                    return setErrorResponse("NVR录像机没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "sxt": [i for i in sxt],
                        "sxt_sl": sxt_number,
                        "fc_zj": [i for i in fc_zj],
                        "fc_zj_sl": sxt_number,
                        "jhj": [i for i in jhj],
                        "jhj_sl": 1,
                        "nvr": [i for i in nvr],
                        "nvr_sl": 1,
                    }
                )
            else:
                return setSuccessResponse(
                    msg="逻辑复杂,脑细包烧光了",
                    data={
                        "提示": "请记录方案需求,联系17366909676",
                        "原因": "sxt 数量大于等于23",
                    }
                )
        # 摄像头为0
        elif sxt_number == 0:
            # 纯 ap 数
            if 1 <= int(ap_number) <= 4:
                # 默认5口一体化
                # 大于 gt  小于： lt 大于等于:gte
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(1, 4)).values("product_kind", "xiaoshoujia")
                # 查询所有字段
                # lyq = LyqInfo.objects.filter(
                #      classify=1, POE__lte=ap_number)
                if not lyq:
                    return setErrorResponse("没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        # 查询所有字段,输出
                        # "lyq": [model_to_dict(i) for i in ls],
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )
            elif 5 <= int(ap_number) <= 8:
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(5, 8)).values("product_kind", "xiaoshoujia")
                if not lyq:
                    return setErrorResponse("AP没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )
            elif 9 <= int(ap_number) <= 12:
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(5, 8)).values("product_kind", "xiaoshoujia")
                if not lyq:
                    return setErrorResponse("路由器没有查询到数据")

                jhj = JhjInfo.objects.filter(
                    classify=4, poe_number__range=(1, 4)).values("product_kind", "xiaoshoujia")

                if not jhj:
                    return setErrorResponse("交换机没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "jhj": [i for i in jhj],
                        "jhj_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )
            elif 13 <= int(ap_number) <= 16:
                lyq = LyqInfo.objects.filter(
                    classify=1, POE__range=(5, 8)).values("product_kind", "xiaoshoujia")
                if not lyq:
                    return setErrorResponse("路由器没有查询到数据")

                jhj = JhjInfo.objects.filter(
                    # 交换出来有三款,判断价格最便宜一款(不会写)
                    classify=4, poe_number__range=(5, 8), upstream_number__gte=1).values("product_kind",
                                                                                         "xiaoshoujia")

                if not jhj:
                    return setErrorResponse("交换机没有查询到数据")
                return setSuccessResponse(
                    msg="ok",
                    data={
                        "lyq": [i for i in lyq],
                        "lyq_sl": 1,
                        "jhj": [i for i in jhj],
                        "jhj_sl": 1,
                        "aps": [i for i in aps],
                        "ap_sl": ap_number,
                    }
                )

        else:
            return setSuccessResponse(
                msg="ok",
                data={
                    "sxt": [i for i in sxt],
                    "sxt_sl": sxt_number,
                    "aps": [i for i in aps],
                    "ap_sl": ap_number,
                }
            )

    except Exception as error:
        traceback.print_exc()
        return setErrorResponse("出错了:" + str(error))


# http://127.0.0.1:8000/api/index/
def indexView(request):
    return render(request, "index.html")


# 动态获取摄像头参数,加了ap
# http://127.0.0.1:8000/api/getSxtOpinions/
def getSxtOpinions(request):
    jiaoju_list = list(SxtInfo.objects.values_list(
        'jiaoju', flat=True).distinct())
    jiaoju_opinions = []
    for i in jiaoju_list:
        jiaoju_opinions.extend(i.split("/"))
    jiaoju1 = list(dict.fromkeys(jiaoju_opinions))

    ap_list = list(ApInfo.objects.values_list(
        'color', flat=True).distinct())
    ap_opinions = []
    for i in ap_list:
        ap_opinions.extend(i.split("/"))
    ap_color = list(dict.fromkeys(ap_opinions))

    return setSuccessResponse(
        msg="ok",
        data={
            # 获取所有不重复的品牌厂商名字列表
            "factory_names": list(SxtInfo.objects.values_list(
                'factory__name', flat=True).distinct()),
            "product_list": [
                {"value": str(type[0]),
                 "label": type[1]}
                for type in SxtInfo.classify_choices],
            "yeshi_choices": [
                {"value": str(type[0]),
                 "label": type[1]}
                for type in SxtInfo.yeshi_choices],
            "pxs": list(SxtInfo.objects.values_list(
                'px', flat=True).distinct()),
            # 焦距
            "jiaoju": jiaoju1,
            # 添加了ap
            "factory_ap": list(ApInfo.objects.values_list(
                'factory__name', flat=True).distinct()),
            "classify_ap": [
                {"value": str(type[0]),
                 "label": type[1]}
                for type in ApInfo.classify_choices],
            "wifi_level_ap": [
                {"value": str(type[0]),
                 "label": type[1]}
                for type in ApInfo.classWiFi_choices],
            "lianjie_sulv_ap": list(ApInfo.objects.values_list(
                'lianjie_sulv', flat=True).distinct()),
            "ap_color": ap_color,

        })


# getGroupAndQuyu
# http://127.0.0.1:8000/api/getGroupAndQuyu/
def getGroupAndQuyu(request):
    return setSuccessResponse(
        msg="ok",
        data={
            "组别": [model_to_dict(i) for i in Group.objects.all()],
            "区域": [model_to_dict(i) for i in Quyu.objects.all()]
        }
    )


# AP 参数查询
def cx_ap(pinpai, feilei: int, sulv: int, shuliang: int):
    result_ap = ApInfo.objects.filter(
        # 查询品牌商的名称 包含关键字
        factory__name__icontains=pinpai,
        classify=feilei,
        lianjie_sulv=sulv,
    ).order_by('xiaoshoujia').first()
    if not result_ap:
        return None
    return [result_ap.to_dict(number=shuliang)]


# 路由器 参数查询
def cx_lyq(pinpai, feilei: int, daijiliang: int, poe_min: int, poe_max: int, shuliang: int):
    result = LyqInfo.objects.filter(
        # 查询品牌商的名称 包含关键字
        factory__name__icontains=pinpai,
        classify=feilei,
        daijialiang__gte=daijiliang,
        POE__range=(poe_min, poe_max)
    ).order_by('xiaoshoujia').first()
    if not result:
        return None
    return [result.to_dict(number=shuliang)]


# 交换机 参数查询
def cx_jhj(feilei: int, poe_min: int, poe_max: int, shuliang: int):
    result = JhjInfo.objects.filter(
        # 查询品牌商的名称 包含关键字
        classify=feilei,
        poe_number__range=(poe_min, poe_max)
    ).order_by('xiaoshoujia').first()
    if not result:
        return None
    return [result.to_dict(number=shuliang)]


# 摄像头 参数查询
def cx_sxt(pinpai, feilei: int, jiaoju, shuliang: int):
    result = SxtInfo.objects.filter(
        # 查询品牌商的名称 包含关键字
        factory__name__icontains=pinpai,
        jiaoju__icontains=jiaoju,
        classify=feilei,

    ).order_by('xiaoshoujia').first()
    if not result:
        return None
    return [result.to_dict(jiaoju, number=shuliang)]


# 录像机 参数查询
def cx_nvr(jierunshu: int, shuliang: int):
    result = NVRInfo.objects.filter(
        # 查询品牌商的名称 包含关键字
        jierunshu=jierunshu,
    ).order_by('xiaoshoujia').first()
    if not result:
        return None
    return [result.to_dict(number=shuliang)]


# 硬盘 参数查询
def cx_yp(classify: int, shuliang: int):
    result = YinPanInfo.objects.filter(
        # 查询品牌商的名称 包含关键字
        classify=classify,
    ).order_by('xiaoshoujia').first()
    if not result:
        return None
    return [result.to_dict(number=shuliang)]


# 辅材 参数查询FucaiInfo
def cx_fc(shuliang: int):
    result = FucaiInfo.objects.filter(
        # 查询品牌商的名称 包含关键字
        classify_fl="枪机支架",
    ).order_by('xiaoshoujia').first()
    if not result:
        return None
    return [result.to_dict(number=shuliang)]


# 综合方案生成
class fascView(APIView):
    def post(self, request, *args, **kwargs):
        str = request.query_params
        print(str)
        jhjtksl = int(str['AP数量']) + int(str['摄像头数量'])
        nvt_lushu = 0
        yp_dx = 0
        poe_min = 0
        poe_max = 0
        # 判断摄像头个数,方便选取录像机
        if 1 <= int(str['摄像头数量']) <= 4:
            nvt_lushu = 4
            yp_dx = 1
        elif 5 <= int(str['摄像头数量']) <= 8:
            nvt_lushu = 8
            yp_dx = 2
        elif 9 <= int(str['摄像头数量']) <= 16:
            nvt_lushu = 16
            yp_dx = 3
        elif 17 <= int(str['摄像头数量']) <= 32:
            nvt_lushu = 32
            yp_dx = 4
        elif 33 <= int(str['摄像头数量']) <= 64:
            nvt_lushu = 64
            yp_dx = 5
        else:
            pass
        nvr = cx_nvr(nvt_lushu, 1)
        yp = cx_yp(yp_dx, 1)
        zj = cx_fc(int(str['摄像头数量']))
        print(nvt_lushu)
        # 判断数量方便下面查询
        if 1 <= int(str['用户数']) <= 100:
            lyq_lx = 1
            if 1 <= int(str['AP数量']) <= 4:
                poe_min = 1
                poe_max = 4
                jhjtksl = jhjtksl - 4
            elif 5 <= int(str['AP数量']) <= 8:
                poe_min = 5
                poe_max = 8
                jhjtksl = jhjtksl - 8
            elif 9 <= int(str['AP数量']) <= 31:
                poe_min = 5
                poe_max = 8
                jhjtksl = jhjtksl - 7

            else:
                pass
        else:
            lyq_lx = 2
        print(jhjtksl)
        jhj_poe_min = 0
        jhj_poe_max = -1
        if 1 <= jhjtksl <= 4:
            jhj_poe_min = 1
            jhj_poe_max = 4
        elif 5 <= jhjtksl <= 8:
            jhj_poe_min = 5
            jhj_poe_max = 8
        elif 9 <= jhjtksl <= 16:
            jhj_poe_min = 9
            jhj_poe_max = 16
        elif 17 <= jhjtksl <= 24:
            jhj_poe_min = 17
            jhj_poe_max = 24
        else:
            pass

        print(jhjtksl, poe_min, poe_max, jhj_poe_min, jhj_poe_max)
        jhjtksl = jhjtksl - 23
        # 第一个交换机
        jhj = cx_jhj(4, jhj_poe_min, jhj_poe_max, 1)
        jhj_poe_min = 0
        jhj_poe_max = -1
        if 1 <= jhjtksl <= 4:
            jhj_poe_min = 1
            jhj_poe_max = 4
        elif 5 <= jhjtksl <= 8:
            jhj_poe_min = 5
            jhj_poe_max = 8
        elif 9 <= jhjtksl <= 16:
            jhj_poe_min = 9
            jhj_poe_max = 16
        elif 17 <= jhjtksl <= 23:
            jhj_poe_min = 17
            jhj_poe_max = 24
        else:
            pass
        jhjtksl = jhjtksl - 22
        # 第二个交换机
        jhj2 = cx_jhj(4, jhj_poe_min, jhj_poe_max, 1)
        jhj_poe_min = 0
        jhj_poe_max = -1
        if 1 <= jhjtksl <= 4:
            jhj_poe_min = 1
            jhj_poe_max = 4
        elif 5 <= jhjtksl <= 8:
            jhj_poe_min = 5
            jhj_poe_max = 8
        elif 9 <= jhjtksl <= 16:
            jhj_poe_min = 9
            jhj_poe_max = 16
        elif 17 <= jhjtksl <= 22:
            jhj_poe_min = 17
            jhj_poe_max = 24
            jhjtksl = jhjtksl - 21

        print(jhjtksl, poe_min, poe_max, jhj_poe_min, jhj_poe_max)
        # 第三个交换机
        jhj3 = cx_jhj(4, jhj_poe_min, jhj_poe_max, 1)
        # 根据数量计算交换机和路由器
        if str['AP数量'] == "0":  # 没有AP就是单摄像头方案
            sxt = cx_sxt(str['摄像头品牌'], 2, str['焦距'], int(str['摄像头数量']))

            return setMyJson(
                msg="ok",
                data={
                    "摄像头": sxt,
                    "枪机支架": zj,
                    "nvr": nvr,
                    "硬盘": yp,
                    "交换机": jhj,
                    "交换机2": jhj2,
                    "交换机3": jhj3,
                })
        elif str['摄像头数量'] == "0":  # 没有摄像头就是单摄像头方案
            ap = cx_ap(str['AP品牌'], 1, int(str['AP速率']), int(str['AP数量']))
            lyq = cx_lyq(str['AP品牌'], lyq_lx, int(str['用户数']), poe_min, poe_max, 1)

            return setMyJson(
                msg="ok",
                data={
                    "AP": ap,
                    "路由器": lyq,
                    "交换机": jhj,
                    "交换机2": jhj2,
                    "交换机3": jhj3,
                })
        else:  # 有AP也有摄像头的方案
            ap = cx_ap(str['AP品牌'], 1, int(str['AP速率']), int(str['AP数量']))
            lyq = cx_lyq(str['AP品牌'], lyq_lx, int(str['用户数']), poe_min, poe_max, 1)
            sxt = cx_sxt(str['摄像头品牌'], 2, str['焦距'], int(str['摄像头数量']))

        # 根据参数写查询逻辑

        return setMyJson(
            msg="ok",
            data={
                "AP": ap,
                "路由器": lyq,
                "交换机": jhj,
                "交换机2": jhj2,
                "交换机3": jhj3,
                "摄像头": sxt,
                "枪机支架": zj,
                "nvr": nvr,
                "硬盘": yp,
            }
        )
